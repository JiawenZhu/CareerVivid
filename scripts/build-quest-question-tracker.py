#!/usr/bin/env python3
"""
Build the CareerVivid quest-question tracker spreadsheet.

Reads (all JSON, no TS parsing):
  - data/interview-guides/<slug>.json          per-company sampleQuestions
  - data/quest-category-banks.json             category taxonomy + tailored banks
  - data/interview-guides/_research-progress.json   audit ledger (which companies done)

Writes:
  - data/quest-question-tracker.xlsx           multi-sheet workbook
  - data/quest-tracker/*.csv                   flat CSV mirrors (git-diff friendly)

The twice-weekly research task runs this after each audit so the table always
reflects the current questions and which companies have been reviewed.

Usage:  python3 scripts/build-quest-question-tracker.py
Requires: openpyxl  (pip install openpyxl)
"""
import json
import csv
import os
import datetime as dt

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
GUIDES_DIR = os.path.join(ROOT, "data", "interview-guides")
BANKS_PATH = os.path.join(ROOT, "data", "quest-category-banks.json")
PROGRESS_PATH = os.path.join(GUIDES_DIR, "_research-progress.json")
XLSX_PATH = os.path.join(ROOT, "data", "quest-question-tracker.xlsx")
CSV_DIR = os.path.join(ROOT, "data", "quest-tracker")

# Round key -> (guide sampleQuestions field OR bank stage, display label)
COMPANY_ROUNDS = [
    ("coding", "Coding"),
    ("systemDesign", "System Design"),
    ("behavioral", "Behavioral"),
    ("values", "Values"),
]
CATEGORY_ROUNDS = [
    ("screening", "Screening"),
    ("behavioral", "Behavioral"),
    ("values", "Values"),
    ("final", "Final"),
]


def load_json(path, default=None):
    if not os.path.exists(path):
        return default
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_guides():
    guides = []
    for fn in sorted(os.listdir(GUIDES_DIR)):
        if not fn.endswith(".json"):
            continue
        if fn in ("_index.json", "_research-progress.json"):
            continue
        d = load_json(os.path.join(GUIDES_DIR, fn))
        if isinstance(d, dict) and d.get("company"):
            guides.append(d)
    guides.sort(key=lambda g: g.get("company", "").lower())
    return guides


def main():
    banks = load_json(BANKS_PATH, {}) or {}
    company_category = banks.get("companyCategory", {})
    question_banks = banks.get("questionBanks", {})
    progress = load_json(PROGRESS_PATH, {}) or {}
    guides = load_guides()

    meta_labels = {
        "big-tech": "Big Tech", "quant-trading": "Quant & Trading", "finance": "Finance & Banking",
        "fintech": "Fintech & Payments", "crypto": "Crypto & Web3", "ai-lab": "AI Labs",
        "dev-infra": "Dev Tools & Infra", "enterprise-saas": "Enterprise SaaS",
        "cybersecurity": "Cybersecurity", "consumer": "Consumer & Media",
        "marketplace": "Marketplace & Commerce", "gaming": "Gaming",
        "hardware": "Hardware & Semiconductors", "deep-tech": "Deep Tech & Defense",
        "healthcare": "Healthcare & Bio",
    }

    # ---- Audit Log rows -------------------------------------------------
    audit_rows = []
    for g in guides:
        slug = g.get("slug", "")
        cat = company_category.get(slug, "(unmapped)")
        sq = g.get("sampleQuestions", {}) or {}
        p = progress.get(slug, {})
        audited = bool(p.get("lastAuditedAt"))
        audit_rows.append({
            "Company": g.get("company", ""),
            "Slug": slug,
            "Category": cat,
            "Category Label": meta_labels.get(cat, cat),
            "Status": "Audited" if audited else "Not yet audited",
            "Last Audited": p.get("lastAuditedAt", ""),
            "Runs": p.get("runs", 0),
            "# Coding": len(sq.get("coding", []) or []),
            "# System Design": len(sq.get("systemDesign", []) or []),
            "# Behavioral": len(sq.get("behavioral", []) or []),
            "# Values": len(sq.get("values", []) or []),
            "Difficulty": g.get("difficulty") if g.get("difficulty") is not None else "",
            "Guide URL": g.get("url", ""),
        })

    # ---- Company Questions rows (all six rounds, exactly what syncs) ----
    # Per-company guide questions PLUS the category-tailored screening/values/
    # final (all difficulty tiers) that get materialized into each company.
    TIERS = ("easy", "medium", "hard")

    def bank_qs(slug, stage, company):
        """Return [(question, tier)] across easy/medium/hard for a stage."""
        cat = company_category.get(slug)
        if not cat:
            return []
        stage_bank = (question_banks.get(cat, {}) or {}).get(stage, {}) or {}
        out = []
        for tier in TIERS:
            for q in stage_bank.get(tier, []) or []:
                out.append((q.replace("{company}", company), tier))
        return out

    company_q_rows = []
    for g in guides:
        slug = g.get("slug", "")
        company = g.get("company", "")
        cat = company_category.get(slug, "(unmapped)")
        sq = g.get("sampleQuestions", {}) or {}

        def add(label, questions, origin):
            """questions: list of (text, difficulty) or plain strings."""
            i = 0
            for item in questions:
                q, diff = item if isinstance(item, tuple) else (item, "")
                if not str(q).strip():
                    continue
                i += 1
                company_q_rows.append({
                    "Company": company, "Slug": slug, "Category": cat,
                    "Round": label, "Source": origin, "Difficulty": diff, "#": i, "Question": q,
                })

        add("Screening", bank_qs(slug, "screening", company), "category")
        add("Coding", sq.get("coding", []) or [], "guide")
        add("System Design", sq.get("systemDesign", []) or [], "guide")
        # behavioral: company-specific first, then difficulty-tiered category bank
        beh = list(sq.get("behavioral", []) or [])
        seen_b = set(beh)
        add("Behavioral", beh, "guide")
        add("Behavioral", [(q, t) for q, t in bank_qs(slug, "behavioral", company) if q not in seen_b], "category")
        # values: company-specific first, then category bank (deduped)
        vals = list(sq.get("values", []) or [])
        seen_v = set(vals)
        add("Values", vals, "guide")
        add("Values", [(q, t) for q, t in bank_qs(slug, "values", company) if q not in seen_v], "category")
        add("Final", bank_qs(slug, "final", company), "category")

    # ---- Category Banks rows (screening / values / final, by tier) -----
    category_q_rows = []
    for cat in sorted(question_banks.keys()):
        for stage, label in CATEGORY_ROUNDS:
            stage_bank = question_banks[cat].get(stage, {}) or {}
            for tier in TIERS:
                for i, q in enumerate(stage_bank.get(tier, []) or [], start=1):
                    category_q_rows.append({
                        "Category": cat,
                        "Category Label": meta_labels.get(cat, cat),
                        "Round": label,
                        "Difficulty": tier,
                        "#": i,
                        "Question ({company} = company name)": q,
                    })

    # ---- Summary rows ---------------------------------------------------
    total = len(guides)
    audited = sum(1 for r in audit_rows if r["Status"] == "Audited")
    by_cat = {}
    for r in audit_rows:
        c = r["Category"]
        by_cat.setdefault(c, {"total": 0, "audited": 0})
        by_cat[c]["total"] += 1
        if r["Status"] == "Audited":
            by_cat[c]["audited"] += 1
    summary_rows = [
        {"Metric": "Generated at", "Value": dt.datetime.now().astimezone().isoformat(timespec="seconds")},
        {"Metric": "Total companies", "Value": total},
        {"Metric": "Audited at least once", "Value": audited},
        {"Metric": "Remaining before full coverage", "Value": total - audited},
        {"Metric": "Coverage", "Value": f"{(audited/total*100):.0f}%" if total else "0%"},
    ]
    cat_summary_rows = [
        {"Category": c, "Category Label": meta_labels.get(c, c),
         "Companies": v["total"], "Audited": v["audited"],
         "Remaining": v["total"] - v["audited"]}
        for c, v in sorted(by_cat.items())
    ]

    write_csvs(audit_rows, company_q_rows, category_q_rows, cat_summary_rows)
    write_xlsx(summary_rows, cat_summary_rows, audit_rows, company_q_rows, category_q_rows)

    print(f"Companies: {total} | audited: {audited} | remaining: {total-audited}")
    print(f"Company questions: {len(company_q_rows)} | category-bank questions: {len(category_q_rows)}")
    print(f"Wrote {XLSX_PATH}")
    print(f"Wrote CSV mirrors in {CSV_DIR}")


def write_csvs(audit_rows, company_q_rows, category_q_rows, cat_summary_rows):
    os.makedirs(CSV_DIR, exist_ok=True)
    def dump(name, rows):
        path = os.path.join(CSV_DIR, name)
        if not rows:
            open(path, "w").close()
            return
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            w.writeheader()
            w.writerows(rows)
    dump("audit-log.csv", audit_rows)
    dump("company-questions.csv", company_q_rows)
    dump("category-banks.csv", category_q_rows)
    dump("category-summary.csv", cat_summary_rows)


def write_xlsx(summary_rows, cat_summary_rows, audit_rows, company_q_rows, category_q_rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    AUDITED_FILL = PatternFill("solid", fgColor="DCFCE7")
    PENDING_FILL = PatternFill("solid", fgColor="FEF3C7")

    wb = Workbook()

    def add_sheet(title, rows, columns, widths, wrap_cols=(), status_col=None):
        ws = wb.create_sheet(title)
        ws.append(columns)
        for ci, _ in enumerate(columns, start=1):
            c = ws.cell(row=1, column=ci)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(vertical="center")
        for r in rows:
            ws.append([r.get(col, "") for col in columns])
        # widths
        for ci, col in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 18)
        # wrap
        for col in wrap_cols:
            if col in columns:
                ci = columns.index(col) + 1
                for ri in range(2, ws.max_row + 1):
                    ws.cell(row=ri, column=ci).alignment = Alignment(wrap_text=True, vertical="top")
        # status colouring
        if status_col and status_col in columns:
            sci = columns.index(status_col) + 1
            for ri in range(2, ws.max_row + 1):
                val = ws.cell(row=ri, column=sci).value
                fill = AUDITED_FILL if val == "Audited" else PENDING_FILL
                ws.cell(row=ri, column=sci).fill = fill
        ws.freeze_panes = "A2"
        if ws.max_row >= 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"
        return ws

    # Summary (first sheet) — replace default
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Value"])
    for ci in (1, 2):
        ws.cell(row=1, column=ci).fill = HEADER_FILL
        ws.cell(row=1, column=ci).font = HEADER_FONT
    for r in summary_rows:
        ws.append([r["Metric"], r["Value"]])
    ws.append([])
    hdr = ["Category", "Category Label", "Companies", "Audited", "Remaining"]
    ws.append(hdr)
    hr = ws.max_row
    for ci in range(1, len(hdr) + 1):
        ws.cell(row=hr, column=ci).fill = HEADER_FILL
        ws.cell(row=hr, column=ci).font = HEADER_FONT
    for r in cat_summary_rows:
        ws.append([r["Category"], r["Category Label"], r["Companies"], r["Audited"], r["Remaining"]])
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 26
    for col in "CDE":
        ws.column_dimensions[col].width = 14

    add_sheet(
        "Audit Log", audit_rows,
        ["Company", "Slug", "Category", "Category Label", "Status", "Last Audited", "Runs",
         "# Coding", "# System Design", "# Behavioral", "# Values", "Difficulty", "Guide URL"],
        {"Company": 22, "Slug": 34, "Category": 16, "Category Label": 22, "Status": 16,
         "Last Audited": 22, "Guide URL": 40},
        status_col="Status",
    )
    add_sheet(
        "Company Questions", company_q_rows,
        ["Company", "Slug", "Category", "Round", "Source", "Difficulty", "#", "Question"],
        {"Company": 22, "Slug": 34, "Category": 16, "Round": 14, "Source": 10,
         "Difficulty": 11, "#": 5, "Question": 104},
        wrap_cols=("Question",),
    )
    add_sheet(
        "Category Banks", category_q_rows,
        ["Category", "Category Label", "Round", "Difficulty", "#", "Question ({company} = company name)"],
        {"Category": 16, "Category Label": 22, "Round": 12, "Difficulty": 11, "#": 5,
         "Question ({company} = company name)": 104},
        wrap_cols=("Question ({company} = company name)",),
    )

    os.makedirs(os.path.dirname(XLSX_PATH), exist_ok=True)
    wb.save(XLSX_PATH)


if __name__ == "__main__":
    main()
